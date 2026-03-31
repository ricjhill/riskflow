"""Generate Excel test spreadsheets for manual testing of the RiskFlow GUI.

Run:  uv run python tests/fixtures/create_test_spreadsheets.py

Creates three files in tests/fixtures/:
  1. reinsurance_bordereaux_messy.xlsx  — standard_reinsurance schema, messy headers
  2. marine_cargo_bordereaux.xlsx       — marine_cargo schema, realistic broker data
  3. multi_sheet_bordereaux.xlsx        — two sheets, one per schema
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

HERE = Path(__file__).parent


def _style_header(ws):
    """Bold white text on dark blue for the header row."""
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")


# ---------------------------------------------------------------------------
# 1. Messy reinsurance bordereaux
#    Uses non-standard column names the SLM hints should resolve:
#      Certificate -> Policy_ID, GWP -> Gross_Premium, TSI -> Sum_Insured, Ccy -> Currency
#    Includes extra columns the mapper should ignore.
# ---------------------------------------------------------------------------
def create_reinsurance():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bordereaux Q1 2025"

    headers = [
        "Certificate No",
        "Insured Name",
        "Risk Location",
        "Effective From",
        "Effective To",
        "TSI (000s)",
        "GWP",
        "Ccy",
        "Broker",
        "Line %",
        "Notes",
    ]
    ws.append(headers)

    rows = [
        [
            "CERT-2025-0001",
            "Acme Industries Ltd",
            "London, UK",
            "01-Jan-2025",
            "31-Dec-2025",
            5000000,
            125000.00,
            "USD",
            "Marsh",
            25.0,
            "Annual renewal — cat XL layer",
        ],
        [
            "CERT-2025-0002",
            "Global Shipping Co",
            "New York, US",
            "15-Feb-2025",
            "14-Feb-2026",
            12000000,
            340000.00,
            "USD",
            "Aon",
            30.0,
            "New business via London market",
        ],
        [
            "CERT-2025-0003",
            "Sakura Holdings KK",
            "Tokyo, JP",
            "01-Mar-2025",
            "28-Feb-2026",
            3500000,
            87500.00,
            "JPY",
            "Willis",
            15.0,
            "",
        ],
        [
            "CERT-2025-0004",
            "Rhine Industrieversicherung",
            "Frankfurt, DE",
            "01-Apr-2025",
            "31-Mar-2026",
            8000000,
            200000.00,
            "EUR",
            "Guy Carpenter",
            50.0,
            "Treaty renewal — pro rata",
        ],
        [
            "CERT-2025-0005",
            "Pacific Re Syndicate",
            "Sydney, AU",
            "10/01/2025",  # DD/MM/YYYY
            "09/01/2026",
            2000000,
            60000.00,
            "USD",
            "Gallagher Re",
            10.0,
            "Retrocession cover",
        ],
        [
            "CERT-2025-0006",
            "Northern Alliance Mutual",
            "Oslo, NO",
            "2025-05-01",  # ISO 8601
            "2026-04-30",
            15000000,
            450000.00,
            "EUR",
            "Marsh",
            40.0,
            "Property excess of loss",
        ],
        [
            "CERT-2025-0007",
            "Cedar Point Underwriters",
            "Toronto, CA",
            "20 March 2025",  # DD Month YYYY
            "19 March 2026",
            6500000,
            195000.00,
            "USD",
            "Aon",
            20.0,
            "Casualty surplus lines",
        ],
        [
            "CERT-2025-0008",
            "Sterling Lloyd's Syndicate 4477",
            "London, UK",
            "01-Jun-2025",
            "31-May-2026",
            25000000,
            750000.00,
            "GBP",
            "Willis",
            35.0,
            "Marine hull — fleet cover",
        ],
        [
            "CERT-2025-0009",
            "Meridian Re Brazil",
            "São Paulo, BR",
            "April 15, 2025",  # Month DD, YYYY
            "April 14, 2026",
            4200000,
            126000.00,
            "USD",
            "Guy Carpenter",
            25.0,
            "Agro risk — crop hail",
        ],
        [
            "CERT-2025-0010",
            "Hong Kong Specialty Re",
            "Hong Kong, HK",
            "2025/07/01",  # YYYY/MM/DD
            "2026/06/30",
            9000000,
            270000.00,
            "USD",
            "Gallagher Re",
            15.0,
            "Political violence",
        ],
    ]
    for r in rows:
        ws.append(r)

    _style_header(ws)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    path = HERE / "reinsurance_bordereaux_messy.xlsx"
    wb.save(path)
    print(f"Created {path}")


# ---------------------------------------------------------------------------
# 2. Marine cargo bordereaux
#    Uses aliases: Ship, Departure, ETA, Sum Insured, Origin, Destination
# ---------------------------------------------------------------------------
def create_marine_cargo():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cargo Declarations"

    headers = [
        "Ref #",
        "Vessel / Ship",
        "Departure Date",
        "ETA",
        "Cargo Description",
        "Sum Insured (USD eq)",
        "Premium",
        "Ccy",
        "Origin Port",
        "Destination Port",
        "Commodity Code",
        "Packaging",
    ]
    ws.append(headers)

    rows = [
        [
            "MC-001",
            "MV Maersk Sealand",
            "2025-03-01",
            "2025-03-18",
            "Electronics — consumer",
            4200000,
            105000,
            "USD",
            "Shenzhen",
            "Rotterdam",
            "8471",
            "Container 40ft",
        ],
        [
            "MC-002",
            "MV CMA CGM Riviera",
            "2025-03-05",
            "2025-03-25",
            "Automotive parts",
            7800000,
            195000,
            "EUR",
            "Hamburg",
            "Shanghai",
            "8708",
            "Break-bulk",
        ],
        [
            "MC-003",
            "MV NYK Vega",
            "2025-03-10",
            "2025-03-28",
            "Frozen seafood",
            1500000,
            45000,
            "JPY",
            "Tokyo",
            "Los Angeles",
            "0303",
            "Reefer container",
        ],
        [
            "MC-004",
            "MV Evergreen Fortune",
            "2025-03-15",
            "2025-04-01",
            "Textiles & garments",
            3200000,
            80000,
            "USD",
            "Ho Chi Minh City",
            "Felixstowe",
            "6204",
            "Container 20ft",
        ],
        [
            "MC-005",
            "MV Yang Ming Unity",
            "2025-03-20",
            "2025-04-05",
            "Chemical precursors",
            9500000,
            285000,
            "SGD",
            "Singapore",
            "Antwerp",
            "2903",
            "ISO tank",
        ],
        [
            "MC-006",
            "MV Hapag Berlin",
            "2025-04-01",
            "2025-04-20",
            "Heavy machinery",
            18000000,
            540000,
            "EUR",
            "Bremerhaven",
            "Santos",
            "8429",
            "Flat rack",
        ],
        [
            "MC-007",
            "MV OOCL Hong Kong",
            "2025-04-05",
            "2025-04-18",
            "Pharmaceuticals",
            6000000,
            180000,
            "HKD",
            "Hong Kong",
            "Sydney",
            "3004",
            "Reefer container",
        ],
        [
            "MC-008",
            "MV MSC Gülsün",
            "2025-04-10",
            "2025-05-02",
            "Wine & spirits",
            2800000,
            84000,
            "GBP",
            "Bordeaux (Le Verdon)",
            "New York",
            "2204",
            "Container 20ft",
        ],
    ]
    for r in rows:
        ws.append(r)

    _style_header(ws)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 24

    path = HERE / "marine_cargo_bordereaux.xlsx"
    wb.save(path)
    print(f"Created {path}")


# ---------------------------------------------------------------------------
# 3. Multi-sheet workbook (one sheet per schema)
# ---------------------------------------------------------------------------
def create_multi_sheet():
    wb = openpyxl.Workbook()

    # Sheet 1: reinsurance-style (clean headers this time)
    ws1 = wb.active
    ws1.title = "Property Treaty"
    ws1.append(
        ["Policy ID", "Inception Date", "Expiry Date", "Sum Insured", "Gross Premium", "Currency"]
    )
    clean_ri = [
        ["PT-2025-001", "2025-01-01", "2025-12-31", 10000000, 300000, "USD"],
        ["PT-2025-002", "2025-02-01", "2026-01-31", 7500000, 225000, "GBP"],
        ["PT-2025-003", "2025-03-01", "2026-02-28", 20000000, 600000, "EUR"],
        ["PT-2025-004", "2025-04-01", "2026-03-31", 4000000, 120000, "JPY"],
        ["PT-2025-005", "2025-05-01", "2026-04-30", 15000000, 450000, "USD"],
    ]
    for r in clean_ri:
        ws1.append(r)
    _style_header(ws1)

    # Sheet 2: marine cargo-style
    ws2 = wb.create_sheet("Marine Facultative")
    ws2.append(
        [
            "Vessel Name",
            "Voyage Date",
            "Arrival Date",
            "Cargo Value",
            "Premium",
            "Currency",
            "Port of Loading",
            "Port of Discharge",
        ]
    )
    clean_mc = [
        ["MV Thor", "2025-06-01", "2025-06-15", 5000000, 125000, "USD", "New Orleans", "Liverpool"],
        ["MV Freya", "2025-06-10", "2025-06-28", 3200000, 96000, "GBP", "Tilbury", "Mumbai"],
        ["MV Odin", "2025-07-01", "2025-07-20", 8000000, 240000, "EUR", "Genoa", "Yokohama"],
    ]
    for r in clean_mc:
        ws2.append(r)
    _style_header(ws2)

    for ws in [ws1, ws2]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 20

    path = HERE / "multi_sheet_bordereaux.xlsx"
    wb.save(path)
    print(f"Created {path}")


if __name__ == "__main__":
    create_reinsurance()
    create_marine_cargo()
    create_multi_sheet()
